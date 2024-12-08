import numpy as np
from sympy import symbols, integrate, sqrt
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

# Initialize Workbook
wb = Workbook()
# 1. Extract points from the provided graph (approximation)
points = [(0, 25), (5, 22), (10, 17), (15, 0)]

# Separate x and y values for the cubic regression model
y_values = [p[0] for p in points]
z_values = [p[1] for p in points]

# 2. Fit a cubic regression model: z = ay^3 + by^2 + cy + d
coefficients = np.polyfit(y_values, z_values, 3)  # a, b, c, d
a, b, c, d = coefficients

# Symbolic function for integration and derivative
y = symbols('y')
z_expr = a * y**3 + b * y**2 + c * y + d
dz_dy_expr = z_expr.diff(y)

# 3. Numerical Integration for Volume
width_x = 50  # Width of the barn in feet
volume_expr = width_x * z_expr
volume_integral = integrate(volume_expr, (y, 0, 15))

# 4. Numerical Integration for Surface Area
surface_expr = sqrt(1 + dz_dy_expr**2)
surface_area_expr = width_x * surface_expr
surface_area_integral = integrate(surface_area_expr, (y, 0, 15))

# Prepare Excel data for better clarity
# Add problem sheet
ws_problem = wb.active
ws_problem.title = "Problema"
ws_problem["A1"] = "Descripción del problema"
ws_problem["A3"] = """
Un ranchero construye un granero de dimensiones de 30 por 50 pies.
La forma del techo es simétrica y se ha elegido un modelo cúbico para describir el perfil del techo.

1. Modelo del perfil del techo: z = ay^3 + by^2 + cy + d
2. Volumen del espacio de almacenaje calculado por integración numérica.
3. Área de la superficie del techo calculada por integración numérica.
"""

# Add coefficients of regression model
ws_problem["A6"] = "Modelo ajustado (z):"
ws_problem["B6"] = f"{a:.4f}y^3 + {b:.4f}y^2 + {c:.4f}y + {d:.4f}"

# Add Volume and Surface Area results
ws_problem["A8"] = "Resultados:"
ws_problem["A9"] = "Volumen del espacio de almacenaje (pies cúbicos):"
ws_problem["B9"] = float(volume_integral)
ws_problem["A10"] = "Área de la superficie del techo (pies cuadrados):"
ws_problem["B10"] = float(surface_area_integral)

# Sheet for Data and Calculations
ws_data = wb.create_sheet("Cálculos")
ws_data.append(["y (pies)", "z (Altura del techo, pies)", "dz/dy (Derivada)"])

# Populate y, z, and dz/dy values in the sheet
dz_dy_values = [dz_dy_expr.subs(y, yi) for yi in y_values]
for yi, zi, dzi in zip(y_values, z_values, dz_dy_values):
    ws_data.append([yi, zi, float(dzi)])

# Add integration explanations
ws_data["E1"] = "Integración para Volumen:"
ws_data["E2"] = f"Volumen = ancho * ∫ z(y) dy | de y=0 a y=15"
ws_data["E3"] = f"Resultado: {float(volume_integral):.4f} pies cúbicos"

ws_data["E5"] = "Integración para Área:"
ws_data["E6"] = f"Área = ancho * ∫ √(1 + (dz/dy)^2) dy | de y=0 a y=15"
ws_data["E7"] = f"Resultado: {float(surface_area_integral):.4f} pies cuadrados"

# Add graph for roof profile
ws_graphs = wb.create_sheet("Gráficas")
chart = ScatterChart()
chart.title = "Perfil del Techo"
chart.x_axis.title = "y (pies)"
chart.y_axis.title = "z (Altura, pies)"

x_values_ref = Reference(ws_data, min_col=1, min_row=2, max_row=len(y_values) + 1)
y_values_ref = Reference(ws_data, min_col=2, min_row=2, max_row=len(z_values) + 1)
series = Series(y_values_ref, x_values_ref, title="Perfil del techo")
chart.series.append(series)
ws_graphs.add_chart(chart, "B2")

# Save the workbook
file_path = "/mnt/data/granero_ingenieria_completo.xlsx"
wb.save(file_path)
file_path
